from datetime import datetime
from django.shortcuts import HttpResponse, render, get_object_or_404
from django.http import HttpRequest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from report.forms import DownloadReportForm
from report.models import Employee


def index(request):
    form = DownloadReportForm()
    return render(request, 'report/index.html', {'form': form})


def download_report(request: HttpRequest) -> HttpResponse:
    employee = get_object_or_404(Employee, id=request.POST.get('employee'))
    reports = employee.work_report.select_related('employee')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response[
        'Content-Disposition'
    ] = 'attachment; filename={date}-{employee}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), employee=employee.name
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'reports'
    columns = [
        'номер заказа покупателя',
        'номер позиции в заказе покупателя',
        'время выполнения(минут)',
        'время заполнения отчета',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 35

    for report in reports:
        row_num += 1
        row = [
            report.order,
            report.item_order,
            report.execution_time,
            report.report_date,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
